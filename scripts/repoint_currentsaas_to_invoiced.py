#!/usr/bin/env python3
"""
Repoint the CurrentSaaS tab in SaaSRevTemplate.xlsx from the plan-rate snapshot
column (Accounts!$BH = Custdatlastsaasamount) to the invoiced-this-period column
(Accounts!$I = SaaSIncomeAmount).

In-place replacement: CurrentSaaS now reports invoiced SaaS instead of run-rate.
Only the CurrentSaaS tab changes; every other tab is left byte-identical.

Usage: python3 scripts/repoint_currentsaas_to_invoiced.py [path-to-xlsx]
"""
import sys
import openpyxl

PATH = sys.argv[1] if len(sys.argv) > 1 else "builder/public/templates/SaaSRevTemplate.xlsx"
TAB = "CurrentSaaS"
OLD_REF = "Accounts!$BH"   # Custdatlastsaasamount (plan-rate snapshot)
NEW_REF = "Accounts!$I"    # SaaSIncomeAmount (invoiced this period)


def formula_inventory(ws):
    inv = {}
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                inv[c.coordinate] = c.value
    return inv


def main():
    wb = openpyxl.load_workbook(PATH, data_only=False)
    if TAB not in wb.sheetnames:
        sys.exit(f"ABORT: tab '{TAB}' not found.")

    before = {name: formula_inventory(wb[name]) for name in wb.sheetnames}
    bh_before = sum(1 for f in before[TAB].values() if OLD_REF in f)
    print(f"'{TAB}': {bh_before} formulas reference {OLD_REF}")

    ws = wb[TAB]
    changed = 0
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and OLD_REF in c.value:
                c.value = c.value.replace(OLD_REF, NEW_REF)
                changed += 1
    print(f"Repointed {changed} formulas {OLD_REF} -> {NEW_REF}")

    wb.save(PATH)
    print(f"Saved {PATH}\n")

    # ---- VERIFY against a fresh load ----
    wb2 = openpyxl.load_workbook(PATH, data_only=False)
    ok = True

    if wb2.sheetnames != list(before.keys()):
        ok = False
        print("FAIL sheet list changed:", wb2.sheetnames)
    else:
        print(f"PASS sheet list unchanged ({len(wb2.sheetnames)} tabs)")

    # Every OTHER tab must be byte-identical
    for name in before:
        if name == TAB:
            continue
        if formula_inventory(wb2[name]) != before[name]:
            ok = False
            print(f"FAIL '{name}' changed unexpectedly")
    else:
        print(f"PASS all non-CurrentSaaS tabs unchanged")

    # CurrentSaaS: 0 BH refs, bh_before I refs, every other formula identical
    after = formula_inventory(wb2[TAB])
    bh = sum(1 for f in after.values() if OLD_REF in f)
    inv = sum(1 for f in after.values() if NEW_REF in f)
    structural = 0
    for coord, f in after.items():
        orig = before[TAB].get(coord)
        if orig is None:
            structural += 1
        elif OLD_REF in orig:
            if f != orig.replace(OLD_REF, NEW_REF):
                structural += 1
        elif f != orig:
            structural += 1
    print(f"CurrentSaaS: {OLD_REF} refs={bh} (want 0), {NEW_REF} refs={inv} "
          f"(want {bh_before}), structural mismatches={structural} (want 0)")
    if bh != 0 or inv < bh_before or structural != 0:
        ok = False
        print("FAIL CurrentSaaS assertions")
    else:
        print("PASS CurrentSaaS assertions")

    print("\n" + ("ALL CHECKS PASSED" if ok else "*** CHECKS FAILED ***"))
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
