#!/usr/bin/env python3
"""
Add a non-destructive "CurrentSaaS (Invoiced)" tab to SaaSRevTemplate.xlsx.

It is an exact twin of the existing CurrentSaaS cohort grid, except the 6 dollar
rows sum the invoiced-this-period column (Accounts!$I = SaaSIncomeAmount) instead
of the plan-rate snapshot column (Accounts!$BH = Custdatlastsaasamount).

Nothing on the original CurrentSaaS tab (or any other tab) is changed.

Usage: python3 scripts/add_currentsaas_invoiced_tab.py [path-to-xlsx]
"""
import sys
import openpyxl

PATH = sys.argv[1] if len(sys.argv) > 1 else "builder/public/templates/SaaSRevTemplate.xlsx"
SRC_TAB = "CurrentSaaS"
NEW_TAB = "CurrentSaaS (Invoiced)"
OLD_REF = "Accounts!$BH"   # Custdatlastsaasamount (plan-rate snapshot)
NEW_REF = "Accounts!$I"    # SaaSIncomeAmount (invoiced this period)


def formula_inventory(ws):
    """coord -> formula string, for every formula cell."""
    inv = {}
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                inv[c.coordinate] = c.value
    return inv


def main():
    wb = openpyxl.load_workbook(PATH, data_only=False)

    if NEW_TAB in wb.sheetnames:
        sys.exit(f"ABORT: tab '{NEW_TAB}' already exists — nothing to do.")
    if SRC_TAB not in wb.sheetnames:
        sys.exit(f"ABORT: source tab '{SRC_TAB}' not found.")

    src = wb[SRC_TAB]

    # Snapshot every original sheet's formula inventory so we can prove nothing
    # other than the new tab changed.
    before = {name: formula_inventory(wb[name]) for name in wb.sheetnames}
    src_bh_before = sum(1 for f in before[SRC_TAB].values() if OLD_REF in f)
    print(f"Source '{SRC_TAB}': {src_bh_before} formulas reference {OLD_REF}")

    # 1) Duplicate the sheet (copies values, formulas, styles, col widths, merges).
    new = wb.copy_worksheet(src)
    new.title = NEW_TAB

    # 2) Place the twin immediately after the original.
    wb.move_sheet(NEW_TAB, offset=-(len(wb.sheetnames) - 1 - (wb.sheetnames.index(SRC_TAB) + 1)))

    # 3) Repoint ONLY the BH (plan-rate) references to I (invoiced) on the twin.
    changed = 0
    for row in new.iter_rows():
        for c in row:
            if isinstance(c.value, str) and OLD_REF in c.value:
                c.value = c.value.replace(OLD_REF, NEW_REF)
                changed += 1
    print(f"Twin '{NEW_TAB}': repointed {changed} formulas {OLD_REF} -> {NEW_REF}")

    wb.save(PATH)
    print(f"Saved {PATH}\n")

    # ---- VERIFY against a fresh load from disk ----
    wb2 = openpyxl.load_workbook(PATH, data_only=False)
    ok = True

    expect_sheets = list(before.keys())
    idx = expect_sheets.index(SRC_TAB)
    expect_sheets.insert(idx + 1, NEW_TAB)
    if wb2.sheetnames != expect_sheets:
        ok = False
        print("FAIL sheet order/list:")
        print("  expected:", expect_sheets)
        print("  got     :", wb2.sheetnames)
    else:
        print(f"PASS sheet list ({len(wb2.sheetnames)} tabs), twin right after original")

    # Originals must be byte-identical in their formulas
    for name in before:
        after = formula_inventory(wb2[name])
        if after != before[name]:
            ok = False
            diff = {k: (before[name].get(k), after.get(k))
                    for k in set(before[name]) | set(after)
                    if before[name].get(k) != after.get(k)}
            print(f"FAIL '{name}' formulas changed ({len(diff)} cells): "
                  f"{list(diff.items())[:3]}")
    else:
        print(f"PASS all {len(before)} original tabs unchanged")

    # Twin assertions
    twin = formula_inventory(wb2[NEW_TAB])
    bh = sum(1 for f in twin.values() if OLD_REF in f)
    inv = sum(1 for f in twin.values() if NEW_REF + ":" in f or f.count(NEW_REF) and NEW_REF in f)
    inv = sum(1 for f in twin.values() if NEW_REF in f)
    # Structure: every non-dollar formula must equal the original CurrentSaaS cell.
    src_inv = before[SRC_TAB]
    structural_mismatch = 0
    for coord, f in twin.items():
        orig = src_inv.get(coord)
        if orig is None:
            structural_mismatch += 1
        elif OLD_REF in orig:
            if f != orig.replace(OLD_REF, NEW_REF):
                structural_mismatch += 1
        elif f != orig:
            structural_mismatch += 1

    print(f"Twin: {OLD_REF} refs = {bh} (want 0), {NEW_REF} refs = {inv} "
          f"(want {src_bh_before}), structural mismatches = {structural_mismatch} (want 0)")
    if bh != 0 or inv < src_bh_before or structural_mismatch != 0:
        ok = False
        print("FAIL twin assertions")
    else:
        print("PASS twin assertions")

    # Merged-cell / dimension parity
    if len(wb2[NEW_TAB].merged_cells.ranges) != len(wb2[SRC_TAB].merged_cells.ranges):
        print(f"WARN merged-cell count differs: twin={len(wb2[NEW_TAB].merged_cells.ranges)} "
              f"src={len(wb2[SRC_TAB].merged_cells.ranges)}")
    else:
        print(f"PASS merged-cell parity ({len(wb2[SRC_TAB].merged_cells.ranges)})")

    print("\n" + ("ALL CHECKS PASSED" if ok else "*** CHECKS FAILED ***"))
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
